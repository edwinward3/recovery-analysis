"""Reads the RT and Companies House files, checks them and prepares them for matching."""

from __future__ import annotations

from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Callable, Iterator
import csv
import hashlib
import io
import re
import zipfile

import numpy as np
import pandas as pd


DAYS_PER_MONTH = 30.44

# RT columns and allowed values

REQUIRED_RT_COLUMNS: tuple[str, ...] = (
    "ID",
    "Date Inserted",
    "JudgmentDate",
    "JudgmentStatus",
    "DefendantType",
    "Jurisdiction",
    "Defendant Company Name",
    "Defendant_Postcode",
)

OPTIONAL_RT_COLUMNS: dict[str, object] = {
    "Amount": pd.NA,
    "Defendant Trading Name": "",
    "Defendant Address": "",
    # Optional RT fields; missing dates are stored as NaT.
    "Satisfaction Date": pd.NaT,
    "Cancellation Date": pd.NaT,
    "Cancellation Reason": "",
    "Status Effective Date": pd.NaT,
    "Snapshot Date": pd.NaT,
}

RT_COLUMNS: tuple[str, ...] = (*REQUIRED_RT_COLUMNS, *OPTIONAL_RT_COLUMNS)

_HEADER_ALIASES: dict[str, str] = {
    "id": "ID",
    "dateinserted": "Date Inserted",
    "judgmentdate": "JudgmentDate",
    "judgmentstatus": "JudgmentStatus",
    "defendanttype": "DefendantType",
    "jurisdiction": "Jurisdiction",
    "defendantcompanyname": "Defendant Company Name",
    "companyname": "Defendant Company Name",
    "defendantpostcode": "Defendant_Postcode",
    "postcode": "Defendant_Postcode",
    "amount": "Amount",
    "defendanttradingname": "Defendant Trading Name",
    "tradingname": "Defendant Trading Name",
    "defendantaddress": "Defendant Address",
    "address": "Defendant Address",
    "satisfactiondate": "Satisfaction Date",
    "datesatisfied": "Satisfaction Date",
    "satisfieddate": "Satisfaction Date",
    "dateofsatisfaction": "Satisfaction Date",
    "cancellationdate": "Cancellation Date",
    "datecancelled": "Cancellation Date",
    "datecanceled": "Cancellation Date",
    "cancelleddate": "Cancellation Date",
    "canceleddate": "Cancellation Date",
    "dateofcancellation": "Cancellation Date",
    "cancellationreason": "Cancellation Reason",
    "reasonforcancellation": "Cancellation Reason",
    "cancelreason": "Cancellation Reason",
    "reasoncancelled": "Cancellation Reason",
    "reasoncanceled": "Cancellation Reason",
    "statuseffectivedate": "Status Effective Date",
    "statusdate": "Status Effective Date",
    "statuschangedate": "Status Effective Date",
    "dateofstatus": "Status Effective Date",
    "snapshotdate": "Snapshot Date",
    "extractdate": "Snapshot Date",
    "asofdate": "Snapshot Date",
    "dataasofdate": "Snapshot Date",
    "observationdate": "Snapshot Date",
}

_OPTIONAL_DATE_COLUMNS: tuple[str, ...] = (
    "Satisfaction Date",
    "Cancellation Date",
    "Status Effective Date",
    "Snapshot Date",
)

# Stop if an unrecognised header may contain outcome or timing data.
_DECISIVE_HEADER_FRAGMENTS: tuple[str, ...] = (
    "satisf",
    "cancel",
    "status",
    "snapshot",
    "asof",
    "effective",
    "history",
    "payment",
    "paid",
    "outcome",
    "recover",
    "settle",
    "remov",
    "setaside",
)

_CODED_VALUE_MAPS: dict[str, dict[str, str]] = {
    "JudgmentStatus": {
        "satisfied": "Satisfied",
        "unsatisfied": "Unsatisfied",
        "cancelled": "Cancelled",
        "canceled": "Cancelled",
    },
    "DefendantType": {
        "corporate": "Corporate",
        "non corporate": "Non-Corporate",
        "noncorporate": "Non-Corporate",
        "consumer": "Consumer",
        "non consumer": "Non-Consumer",
        "nonconsumer": "Non-Consumer",
    },
    "Jurisdiction": {
        "england and wales": "England and Wales",
        "england wales": "England and Wales",
        "scotland": "Scotland",
    },
}


# Information recorded about the RT file

@dataclass(frozen=True, slots=True)
class DataAudit:
    """Facts about how the RT file was read; unknown extra headings stay inside RT."""

    rows: int
    observation_date: str
    date_inserted_distinct: int
    judgment_date_min: str
    judgment_date_max: str
    date_inserted_minus_judgment_days_min: int
    date_inserted_minus_judgment_days_median: float
    date_inserted_minus_judgment_days_max: int
    date_inserted_before_judgment_rows: int
    date_inserted_after_observation_rows: int
    age_at_observation_months_min: float
    age_at_observation_months_median: float
    age_at_observation_months_max: float
    invalid_amount_rows: int
    missing_company_name_rows: int
    missing_postcode_rows: int
    data_construct: str
    event_date_columns_present: tuple[str, ...]
    satisfaction_date_present_rows: int
    cancellation_date_present_rows: int
    cancellation_reason_present_rows: int
    status_effective_date_present_rows: int
    snapshot_date_present_rows: int
    historical_snapshots_available: bool
    raw_headers: tuple[str, ...]
    raw_header_schema: tuple[tuple[str, str], ...]
    extra_headers: tuple[str, ...]
    absent_optional_columns: tuple[str, ...]
    raw_source_sha256: str
    raw_header_schema_sha256: str
    analysis_fingerprint: str
    provenance_fingerprint: str
    status_counts: dict[str, int]
    defendant_type_counts: dict[str, int]
    jurisdiction_counts: dict[str, int]

    def as_dict(self) -> dict[str, object]:
        return asdict(self)


# Read and check the RT file

def _canon_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def _canon_code(value: object) -> str:
    text = str(value).strip().lower().replace("&", " and ")
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def _observation_timestamp(value: str | pd.Timestamp) -> pd.Timestamp:
    if isinstance(value, str) and re.fullmatch(r"\d{4}-\d{2}-\d{2}", value) is None:
        raise ValueError("RT extract date must use YYYY-MM-DD")
    try:
        if isinstance(value, str):
            stamp = pd.to_datetime(value, format="%Y-%m-%d", errors="raise")
        else:
            stamp = pd.Timestamp(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"invalid RT extract date: {value!r}") from exc
    if pd.isna(stamp):
        raise ValueError("RT extract date is missing")
    if stamp.tzinfo is not None:
        stamp = stamp.tz_convert(None)
    return stamp.normalize()


def _reject_duplicate_raw_headers(headers: tuple[object, ...], source: str) -> None:
    seen: dict[str, str] = {}
    duplicates: list[str] = []
    for value in headers:
        display = "" if value is None else str(value).strip()
        canonical = display.casefold()
        if canonical in seen:
            duplicate = display or "<blank>"
            if duplicate not in duplicates:
                duplicates.append(duplicate)
        else:
            seen[canonical] = display
    if duplicates:
        raise ValueError(f"{source} has duplicate raw header(s): {duplicates}")


def _csv_headers(path: Path, encoding: str, *, errors: str = "strict") -> tuple[str, ...]:
    with path.open("r", encoding=encoding, errors=errors, newline="") as handle:
        return tuple(next(csv.reader(handle), ()))


def _read_csv(path: Path) -> pd.DataFrame:
    problem: UnicodeDecodeError | None = None
    for encoding in ("utf-8-sig", "cp1252"):
        try:
            _reject_duplicate_raw_headers(
                _csv_headers(path, encoding),
                path.name,
            )
            return pd.read_csv(
                path,
                dtype=str,
                encoding=encoding,
                keep_default_na=False,
                low_memory=False,
            )
        except UnicodeDecodeError as exc:
            problem = exc
    raise ValueError(
        f"{path.name} is not readable as UTF-8 or Windows-1252; re-save it as CSV UTF-8"
    ) from problem


def _read_rt_file(path: Path) -> pd.DataFrame:
    suffix = path.suffix.lower()
    try:
        if suffix == ".csv":
            return _read_csv(path)
        if suffix in {".xlsx", ".xlsm"}:
            try:
                import openpyxl

                workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
                try:
                    if len(workbook.sheetnames) != 1:
                        raise ValueError(
                            "RT workbook must contain exactly one worksheet; "
                            f"found {len(workbook.sheetnames)}"
                        )
                    if workbook.active.sheet_state != "visible":
                        raise ValueError("RT workbook worksheet must be visible")
                    headers = tuple(
                        next(
                            workbook.active.iter_rows(
                                min_row=1,
                                max_row=1,
                                values_only=True,
                            ),
                            (),
                        )
                    )
                finally:
                    workbook.close()
                _reject_duplicate_raw_headers(headers, path.name)
                return pd.read_excel(path, dtype=str, engine="openpyxl")
            except ImportError as exc:
                raise ValueError("XLSX input requires the bundled openpyxl package") from exc
    except Exception as exc:
        raise ValueError(f"could not read {path.name}: {type(exc).__name__}: {exc}") from exc
    raise ValueError(f"unsupported RT input format {suffix!r}; expected .csv, .xlsx, or .xlsm")


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _sequence_fingerprint(values: tuple[tuple[str, str], ...]) -> str:
    """Hash an ordered sequence without ambiguous string concatenation."""

    digest = hashlib.sha256()
    for original, standardised in values:
        for value in (original, standardised):
            encoded = value.encode("utf-8")
            digest.update(len(encoded).to_bytes(8, "big"))
            digest.update(encoded)
    return digest.hexdigest()


def _provenance_fingerprint(
    *,
    raw_source_sha256: str,
    raw_header_schema_sha256: str,
    analysis_fingerprint: str,
) -> str:
    digest = hashlib.sha256()
    for label, value in (
        ("raw_source_sha256", raw_source_sha256),
        ("raw_header_schema_sha256", raw_header_schema_sha256),
        ("analysis_fingerprint", analysis_fingerprint),
    ):
        digest.update(label.encode("ascii"))
        digest.update(b"\x00")
        digest.update(value.encode("ascii"))
        digest.update(b"\x00")
    return digest.hexdigest()


def _standardise_headers(frame: pd.DataFrame) -> pd.DataFrame:
    rename: dict[object, str] = {}
    seen: dict[str, object] = {}
    raw_headers = tuple(str(column) for column in frame.columns)
    raw_header_schema: list[tuple[str, str]] = []
    decisive_unknown: list[str] = []
    for original in frame.columns:
        canonical = _canon_header(original)
        expected = _HEADER_ALIASES.get(canonical)
        raw_header_schema.append((str(original), expected or "<unrecognised>"))
        if expected is None:
            if any(fragment in canonical for fragment in _DECISIVE_HEADER_FRAGMENTS):
                decisive_unknown.append(str(original))
            continue
        if expected in seen:
            raise ValueError(
                f"duplicate columns map to {expected!r}: {seen[expected]!r} and {original!r}"
            )
        seen[expected] = original
        rename[original] = expected
    if decisive_unknown:
        raise ValueError(
            "RT extract has unrecognised outcome/history-related column(s): "
            f"{decisive_unknown}; add an explicit reviewed header alias before use"
        )
    missing = [column for column in REQUIRED_RT_COLUMNS if column not in seen]
    if missing:
        raise ValueError(f"RT extract is missing required column(s): {missing}")
    result = frame.rename(columns=rename).copy()
    for column, default in OPTIONAL_RT_COLUMNS.items():
        if column not in result:
            result[column] = default
    extra_headers = tuple(str(column) for column in result.columns if column not in RT_COLUMNS)
    absent_optional = tuple(column for column in OPTIONAL_RT_COLUMNS if column not in seen)
    # Extra fields have been read for the audit; drop them before later steps.
    result = result.loc[:, list(RT_COLUMNS)].copy()
    result.attrs["extra_headers"] = extra_headers
    result.attrs["absent_optional"] = absent_optional
    result.attrs["raw_headers"] = raw_headers
    result.attrs["raw_header_schema"] = tuple(raw_header_schema)
    return result


def _normalise_coded_columns(frame: pd.DataFrame) -> None:
    for column, mapping in _CODED_VALUE_MAPS.items():
        raw = frame[column].astype("string").fillna("").str.strip()
        canonical = raw.map(lambda value: mapping.get(_canon_code(value), pd.NA))
        invalid = canonical.isna()
        if invalid.any():
            raise ValueError(
                f"{column} has {int(invalid.sum())} missing or unsupported row(s) "
                f"across {int(raw[invalid].nunique(dropna=False))} distinct value(s)"
            )
        frame[column] = canonical.astype(str)


def _parse_required_date(frame: pd.DataFrame, column: str) -> pd.Series:
    raw = frame[column].astype("string").fillna("").str.strip()
    iso = raw.str.fullmatch(r"\d{4}-\d{2}-\d{2}(?:[ T].*)?", na=False)
    parsed = pd.Series(pd.NaT, index=frame.index, dtype="datetime64[ns]")
    if iso.any():
        parsed.loc[iso] = pd.to_datetime(
            raw.loc[iso], format="ISO8601", errors="coerce"
        ).astype("datetime64[ns]")
    if (~iso).any():
        parsed.loc[~iso] = pd.to_datetime(
            raw.loc[~iso], format="mixed", dayfirst=True, errors="coerce"
        ).astype("datetime64[ns]")
    invalid = parsed.isna()
    if invalid.any():
        raise ValueError(f"{column} has {int(invalid.sum())} missing or unparseable row(s)")
    return parsed.dt.normalize()


def _parse_optional_date(frame: pd.DataFrame, column: str) -> pd.Series:
    raw = frame[column].astype("string").fillna("").str.strip()
    populated = raw.ne("")
    parsed = pd.Series(pd.NaT, index=frame.index, dtype="datetime64[ns]")
    if populated.any():
        iso = raw.str.fullmatch(r"\d{4}-\d{2}-\d{2}(?:[ T].*)?", na=False)
        iso_populated = populated & iso
        other_populated = populated & ~iso
        if iso_populated.any():
            parsed.loc[iso_populated] = pd.to_datetime(
                raw.loc[iso_populated], format="ISO8601", errors="coerce"
            ).astype("datetime64[ns]")
        if other_populated.any():
            parsed.loc[other_populated] = pd.to_datetime(
                raw.loc[other_populated], format="mixed", dayfirst=True, errors="coerce"
            ).astype("datetime64[ns]")
    invalid = populated & parsed.isna()
    if invalid.any():
        raise ValueError(f"{column} has {int(invalid.sum())} unparseable row(s)")
    return parsed.dt.normalize()


def _validate_optional_timing(
    frame: pd.DataFrame,
    observed: pd.Timestamp,
    present_optional: set[str],
) -> None:
    """Check supplied event fields against dates and judgment status."""

    for column in _OPTIONAL_DATE_COLUMNS:
        if column not in present_optional:
            continue
        before_judgment = frame[column].notna() & (
            frame[column] < frame["JudgmentDate"]
        )
        after_observation = frame[column].notna() & (frame[column] > observed)
        if before_judgment.any():
            raise ValueError(
                f"{column} is before JudgmentDate for "
                f"{int(before_judgment.sum())} row(s)"
            )
        if after_observation.any():
            raise ValueError(
                f"{column} is after the RT extract date for "
                f"{int(after_observation.sum())} row(s)"
            )

    status = frame["JudgmentStatus"]
    if "Satisfaction Date" in present_optional:
        satisfied_without_date = status.eq("Satisfied") & frame["Satisfaction Date"].isna()
        unsatisfied_with_date = status.eq("Unsatisfied") & frame["Satisfaction Date"].notna()
        if satisfied_without_date.any():
            raise ValueError(
                "Satisfaction Date is present in the schema but missing for "
                f"{int(satisfied_without_date.sum())} Satisfied row(s)"
            )
        if unsatisfied_with_date.any():
            raise ValueError(
                "Satisfaction Date is populated for "
                f"{int(unsatisfied_with_date.sum())} Unsatisfied row(s)"
            )

    if "Cancellation Date" in present_optional:
        cancelled_without_date = status.eq("Cancelled") & frame["Cancellation Date"].isna()
        other_with_date = ~status.eq("Cancelled") & frame["Cancellation Date"].notna()
        if cancelled_without_date.any():
            raise ValueError(
                "Cancellation Date is present in the schema but missing for "
                f"{int(cancelled_without_date.sum())} Cancelled row(s)"
            )
        if other_with_date.any():
            raise ValueError(
                "Cancellation Date is populated for "
                f"{int(other_with_date.sum())} non-Cancelled row(s)"
            )

    if "Cancellation Reason" in present_optional:
        reason_present = frame["Cancellation Reason"].ne("")
        other_with_reason = ~status.eq("Cancelled") & reason_present
        if other_with_reason.any():
            raise ValueError(
                "Cancellation Reason is populated for "
                f"{int(other_with_reason.sum())} non-Cancelled row(s)"
            )

    both_events = frame["Satisfaction Date"].notna() & frame["Cancellation Date"].notna()
    cancellation_before_satisfaction = both_events & (
        frame["Cancellation Date"] < frame["Satisfaction Date"]
    )
    if cancellation_before_satisfaction.any():
        raise ValueError(
            "Cancellation Date is before Satisfaction Date for "
            f"{int(cancellation_before_satisfaction.sum())} row(s)"
        )

    if "Status Effective Date" in present_optional:
        satisfied_order = (
            status.eq("Satisfied")
            & frame["Status Effective Date"].notna()
            & frame["Satisfaction Date"].notna()
            & (frame["Status Effective Date"] < frame["Satisfaction Date"])
        )
        cancelled_order = (
            status.eq("Cancelled")
            & frame["Status Effective Date"].notna()
            & frame["Cancellation Date"].notna()
            & (frame["Status Effective Date"] < frame["Cancellation Date"])
        )
        if satisfied_order.any() or cancelled_order.any():
            raise ValueError(
                "Status Effective Date precedes the corresponding event date for "
                f"{int((satisfied_order | cancelled_order).sum())} row(s)"
            )

    if "Snapshot Date" in present_optional:
        missing = frame["Snapshot Date"].isna()
        if missing.any():
            raise ValueError(
                "Snapshot Date is present in the schema but missing for "
                f"{int(missing.sum())} row(s)"
            )
        distinct = int(frame["Snapshot Date"].nunique())
        if distinct != 1:
            raise ValueError(
                f"Snapshot Date has {distinct} distinct values; a cross-sectional "
                "extract must have exactly one"
            )
        snapshot = frame["Snapshot Date"].iloc[0]
        if snapshot != observed:
            raise ValueError(
                f"Snapshot Date {snapshot.date().isoformat()} does not match "
                f"RT extract date {observed.date().isoformat()}"
            )


def _classify_data_construct(present_optional: set[str]) -> str:
    event_columns = present_optional.intersection(
        {"Satisfaction Date", "Cancellation Date", "Status Effective Date"}
    )
    snapshot_present = "Snapshot Date" in present_optional
    reason_present = "Cancellation Reason" in present_optional
    if event_columns and snapshot_present:
        return "status_with_event_dates_and_snapshot_date"
    if event_columns:
        return "status_with_event_dates_unique_judgment_rows"
    if reason_present and snapshot_present:
        return "status_with_cancellation_reason_and_snapshot_date"
    if reason_present:
        return "status_with_cancellation_reason_unique_judgment_rows"
    if not snapshot_present:
        return "status_only_unique_judgment_rows"
    return "status_with_snapshot_date_unique_judgment_rows"


def _parse_amount(series: pd.Series) -> tuple[pd.Series, int]:
    raw = series.astype("string").fillna("").str.strip()
    cleaned = (
        raw.str.replace(",", "", regex=False)
        .str.replace("£", "", regex=False)
        .str.replace(r"^\((.*)\)$", r"-\1", regex=True)
    )
    values = pd.to_numeric(cleaned.where(cleaned.ne("")), errors="coerce")
    invalid = raw.ne("") & (values.isna() | ~np.isfinite(values) | values.lt(0))
    values = values.mask(invalid)
    return values.astype(float), int(invalid.sum())


def frame_fingerprint(frame: pd.DataFrame) -> str:
    """Fingerprint a pinned-pandas frame without rereading its source file."""

    digest = hashlib.sha256()
    digest.update("\x1f".join(map(str, frame.columns)).encode("utf-8"))
    row_hashes = pd.util.hash_pandas_object(frame, index=False, categorize=False)
    digest.update(row_hashes.to_numpy(dtype="uint64", copy=False).tobytes())
    return digest.hexdigest()


def read_rt_extract(
    path: str | Path,
    observation_date: str | pd.Timestamp,
) -> tuple[pd.DataFrame, DataAudit]:
    """Read, standardise, and validate an RT extract.

    The returned frame preserves one row per judgment and adds only explicit
    timing fields.  Invalid identifiers, coded values, or dates stop the run.
    """

    source = Path(path)
    if not source.is_file():
        raise ValueError(f"RT extract does not exist: {source}")
    observed = _observation_timestamp(observation_date)
    raw_source_sha256 = _file_sha256(source)
    frame = _standardise_headers(_read_rt_file(source))
    extra_headers = tuple(frame.attrs.get("extra_headers", ()))
    absent_optional = tuple(frame.attrs.get("absent_optional", ()))
    raw_headers = tuple(frame.attrs.get("raw_headers", ()))
    raw_header_schema = tuple(frame.attrs.get("raw_header_schema", ()))
    raw_header_schema_sha256 = _sequence_fingerprint(raw_header_schema)
    present_optional = set(OPTIONAL_RT_COLUMNS).difference(absent_optional)
    if frame.empty:
        raise ValueError("RT extract contains no judgment rows")

    ids = frame["ID"].astype("string").fillna("").str.strip()
    if ids.eq("").any():
        raise ValueError(f"ID has {int(ids.eq('').sum())} missing row(s)")
    duplicated = ids.duplicated(keep=False)
    if duplicated.any():
        if "Snapshot Date" in present_optional:
            raise ValueError(
                "repeated judgment IDs with Snapshot Date indicate historical "
                "snapshots; the one-row-per-judgment reader cannot analyse that "
                "structure"
            )
        raise ValueError(
            f"ID must be unique; {int(duplicated.sum())} row(s) share duplicate ID values"
        )
    frame["ID"] = ids.astype(str)

    _normalise_coded_columns(frame)
    frame["Date Inserted"] = _parse_required_date(frame, "Date Inserted")
    frame["JudgmentDate"] = _parse_required_date(frame, "JudgmentDate")
    for column in _OPTIONAL_DATE_COLUMNS:
        frame[column] = _parse_optional_date(frame, column)

    insertion_before_judgment = frame["Date Inserted"] < frame["JudgmentDate"]
    judgment_after_observation = frame["JudgmentDate"] > observed
    insertion_after_observation = frame["Date Inserted"] > observed
    if judgment_after_observation.any():
        raise ValueError(
            f"JudgmentDate is after the RT extract date for "
            f"{int(judgment_after_observation.sum())} row(s)"
        )

    for column in (
        "Defendant Company Name",
        "Defendant Trading Name",
        "Defendant Address",
        "Defendant_Postcode",
    ):
        frame[column] = frame[column].astype("string").fillna("").str.strip().astype(str)
    frame["Cancellation Reason"] = (
        frame["Cancellation Reason"].astype("string").fillna("").str.strip().astype(str)
    )

    _validate_optional_timing(frame, observed, present_optional)

    amounts, invalid_amounts = _parse_amount(frame["Amount"])
    frame["Amount"] = amounts
    inserted_minus_judgment = (
        frame["Date Inserted"] - frame["JudgmentDate"]
    ).dt.days
    age_days = (observed - frame["JudgmentDate"]).dt.days
    frame["date_inserted_minus_judgment_days"] = inserted_minus_judgment.astype(int)
    frame["age_at_observation_days"] = age_days.astype(int)
    frame["age_at_observation_months"] = age_days / DAYS_PER_MONTH
    frame.attrs["observation_date"] = observed.date().isoformat()
    frame.attrs["data_construct"] = _classify_data_construct(present_optional)
    analysis_fingerprint = frame_fingerprint(frame)
    provenance_fingerprint = _provenance_fingerprint(
        raw_source_sha256=raw_source_sha256,
        raw_header_schema_sha256=raw_header_schema_sha256,
        analysis_fingerprint=analysis_fingerprint,
    )

    audit = DataAudit(
        rows=int(len(frame)),
        observation_date=observed.date().isoformat(),
        date_inserted_distinct=int(frame["Date Inserted"].nunique()),
        judgment_date_min=frame["JudgmentDate"].min().date().isoformat(),
        judgment_date_max=frame["JudgmentDate"].max().date().isoformat(),
        date_inserted_minus_judgment_days_min=int(inserted_minus_judgment.min()),
        date_inserted_minus_judgment_days_median=float(inserted_minus_judgment.median()),
        date_inserted_minus_judgment_days_max=int(inserted_minus_judgment.max()),
        date_inserted_before_judgment_rows=int(insertion_before_judgment.sum()),
        date_inserted_after_observation_rows=int(insertion_after_observation.sum()),
        age_at_observation_months_min=float(frame["age_at_observation_months"].min()),
        age_at_observation_months_median=float(frame["age_at_observation_months"].median()),
        age_at_observation_months_max=float(frame["age_at_observation_months"].max()),
        invalid_amount_rows=invalid_amounts,
        missing_company_name_rows=int(frame["Defendant Company Name"].eq("").sum()),
        missing_postcode_rows=int(frame["Defendant_Postcode"].eq("").sum()),
        data_construct=str(frame.attrs["data_construct"]),
        event_date_columns_present=tuple(
            column
            for column in _OPTIONAL_DATE_COLUMNS[:-1]
            if column in present_optional
        ),
        satisfaction_date_present_rows=int(frame["Satisfaction Date"].notna().sum()),
        cancellation_date_present_rows=int(frame["Cancellation Date"].notna().sum()),
        cancellation_reason_present_rows=int(frame["Cancellation Reason"].ne("").sum()),
        status_effective_date_present_rows=int(
            frame["Status Effective Date"].notna().sum()
        ),
        snapshot_date_present_rows=int(frame["Snapshot Date"].notna().sum()),
        # One extract cannot show status history for a judgment.
        historical_snapshots_available=False,
        raw_headers=raw_headers,
        raw_header_schema=raw_header_schema,
        extra_headers=extra_headers,
        absent_optional_columns=absent_optional,
        raw_source_sha256=raw_source_sha256,
        raw_header_schema_sha256=raw_header_schema_sha256,
        analysis_fingerprint=analysis_fingerprint,
        provenance_fingerprint=provenance_fingerprint,
        status_counts={
            str(key): int(value) for key, value in frame["JudgmentStatus"].value_counts().items()
        },
        defendant_type_counts={
            str(key): int(value) for key, value in frame["DefendantType"].value_counts().items()
        },
        jurisdiction_counts={
            str(key): int(value) for key, value in frame["Jurisdiction"].value_counts().items()
        },
    )
    return frame, audit


# Read the Companies House file in chunks

def iter_ch_chunks(
    path: str | Path,
    *,
    chunksize: int = 100_000,
    usecols: Callable[[object], bool] | None = None,
) -> Iterator[pd.DataFrame]:
    """Yield one CH CSV, directly or as the sole CSV in a ZIP, in chunks."""

    source = Path(path)
    if not source.is_file():
        raise ValueError(f"Companies House input does not exist: {source}")
    if chunksize <= 0:
        raise ValueError("chunksize must be positive")
    read_options: dict[str, object] = {
        "dtype": str,
        "chunksize": chunksize,
        "keep_default_na": False,
        "encoding": "utf-8-sig",
        "encoding_errors": "replace",
    }
    if usecols is not None:
        read_options["usecols"] = usecols

    if source.suffix.lower() == ".zip":
        try:
            with zipfile.ZipFile(source) as archive:
                csv_names = sorted(
                    name for name in archive.namelist() if name.lower().endswith(".csv")
                )
                if not csv_names:
                    raise ValueError(f"{source.name} contains no CSV file")
                if len(csv_names) != 1:
                    raise ValueError(
                        f"{source.name} must contain exactly one CSV file; "
                        f"found {len(csv_names)}"
                    )
                with archive.open(csv_names[0]) as header_handle:
                    with io.TextIOWrapper(
                        header_handle,
                        encoding="utf-8-sig",
                        errors="replace",
                        newline="",
                    ) as text_handle:
                        headers = tuple(next(csv.reader(text_handle), ()))
                _reject_duplicate_raw_headers(
                    headers,
                    f"{source.name}:{csv_names[0]}",
                )
                with archive.open(csv_names[0]) as handle:
                    yield from pd.read_csv(handle, **read_options)
        except zipfile.BadZipFile as exc:
            raise ValueError(f"{source.name} is not a readable ZIP file") from exc
        return
    if source.suffix.lower() != ".csv":
        raise ValueError("Companies House input must be a .csv or .zip")
    _reject_duplicate_raw_headers(
        _csv_headers(source, "utf-8-sig", errors="replace"),
        source.name,
    )
    yield from pd.read_csv(source, **read_options)


__all__ = [
    "DAYS_PER_MONTH",
    "DataAudit",
    "OPTIONAL_RT_COLUMNS",
    "REQUIRED_RT_COLUMNS",
    "RT_COLUMNS",
    "frame_fingerprint",
    "iter_ch_chunks",
    "read_rt_extract",
]
